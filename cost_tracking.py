"""
Cost Tracking Module
Maliyet verilerindeki değişiklikleri Excel'e kaydetme
"""

import os
import pandas as pd
from datetime import datetime
import logging
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_yearly_excel_filename():
    """Yıllık Excel dosya adını döndür"""
    current_year = datetime.now().year
    return f"cost_changes_{current_year}.xlsx"


def get_current_turkey_time():
    """Türkiye saatini döndür"""
    turkey_tz = pytz.timezone('Europe/Istanbul')
    now = datetime.now(turkey_tz)
    return now.strftime("%d.%m.%Y %H:%M:%S")


def create_cost_tracking_excel():
    """Excel dosyası yoksa oluştur ve başlıkları ekle"""
    filename = get_yearly_excel_filename()
    
    if os.path.exists(filename):
        return filename
    
    try:
        # Yeni workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Maliyet Değişiklikleri"
        
        # Başlık satırı oluştur
        headers = [
            # Temel bilgiler
            "Tarih-Saat", "Kullanıcı", "Barkod", "Ürün Adı", "Satış Fiyatı",
            
            # Üretim giderleri (20 satır)
            "Üretim_1_İsim", "Üretim_1_Tutar",
            "Üretim_2_İsim", "Üretim_2_Tutar",
            "Üretim_3_İsim", "Üretim_3_Tutar",
            "Üretim_4_İsim", "Üretim_4_Tutar",
            "Üretim_5_İsim", "Üretim_5_Tutar",
            "Üretim_6_İsim", "Üretim_6_Tutar",
            "Üretim_7_İsim", "Üretim_7_Tutar",
            "Üretim_8_İsim", "Üretim_8_Tutar",
            "Üretim_9_İsim", "Üretim_9_Tutar",
            "Üretim_10_İsim", "Üretim_10_Tutar",
            "Üretim_11_İsim", "Üretim_11_Tutar",
            "Üretim_12_İsim", "Üretim_12_Tutar",
            "Üretim_13_İsim", "Üretim_13_Tutar",
            "Üretim_14_İsim", "Üretim_14_Tutar",
            "Üretim_15_İsim", "Üretim_15_Tutar",
            "Üretim_16_İsim", "Üretim_16_Tutar",
            "Üretim_17_İsim", "Üretim_17_Tutar",
            "Üretim_18_İsim", "Üretim_18_Tutar",
            "Üretim_19_İsim", "Üretim_19_Tutar",
            "Üretim_20_İsim", "Üretim_20_Tutar",
            
            # Üretim toplam KDV
            "Üretim_Toplam_KDV",
            
            # Diğer giderler ve hesaplamalar
            "Kargo_Ücreti", "Kargo_KDV", 
            "Komisyon_Tutarı", "Komisyon_KDV",
            "Stopaj_Tutarı", "Platform_Bedeli", "Diğer_Gider_Tutarı",
            "Hesaplanan_KDV", "Toplam_Giderler", "Net_KDV_Yükümlülüğü", 
            "Kar_Tutarı", "Kar_Oranı"
        ]
        
        # Başlıkları ekle
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 18  # Tarih-Saat
        ws.column_dimensions['B'].width = 12  # Kullanıcı
        ws.column_dimensions['C'].width = 15  # Barkod
        ws.column_dimensions['D'].width = 25  # Ürün Adı
        
        # Diğer sütunlar için genel genişlik
        for col in range(5, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
        
        # Dosyayı kaydet
        wb.save(filename)
        logging.info(f"Yeni maliyet takip Excel dosyası oluşturuldu: {filename}")
        
        return filename
        
    except Exception as e:
        logging.error(f"Excel dosyası oluşturma hatası: {str(e)}")
        return None


def log_cost_data_change(barcode, product_title, username, cost_data, profit_analysis):
    """Maliyet verisi değişikliğini Excel'e kaydet"""
    try:
        filename = create_cost_tracking_excel()
        if not filename:
            return False
        
        # Mevcut Excel'i aç
        wb = load_workbook(filename)
        ws = wb.active
        
        # Yeni satır numarasını bul
        next_row = ws.max_row + 1
        
        # Temel bilgiler
        turkey_time = get_current_turkey_time()
        sale_price = float(cost_data.get('sale_price', 0))
        
        # Satırda yazılacak veriler
        row_data = [
            turkey_time,
            username,
            barcode,
            product_title[:50] if product_title else '',  # Başlığı kısalt
            sale_price
        ]
        
        # Üretim giderleri (20 satır)
        production_costs = cost_data.get('production_costs', [])
        production_total_vat = 0  # Üretim toplam KDV
        
        for i in range(20):
            if i < len(production_costs):
                cost_item = production_costs[i]
                amount = float(cost_item.get('amount', 0))
                production_total_vat += amount * 0.2  # KDV topla
                row_data.extend([
                    cost_item.get('name', ''),
                    amount
                ])
            else:
                row_data.extend(['', 0])  # Boş satırlar
        
        # Üretim toplam KDV'yi ekle
        row_data.append(production_total_vat)
        
        # Diğer giderler ve hesaplamalar
        cargo_cost = float(cost_data.get('cargo_cost', 0))
        cargo_vat = cargo_cost - (cargo_cost / 1.2) if cargo_cost > 0 else 0
        
        commission_rate = float(cost_data.get('commission_rate', 0))
        commission_amount = sale_price * (commission_rate / 100) if commission_rate > 0 else 0
        commission_vat = commission_amount - (commission_amount / 1.2) if commission_amount > 0 else 0
        
        withholding_rate = float(cost_data.get('withholding_rate', 0))
        withholding_amount = sale_price * (withholding_rate / 100) if withholding_rate > 0 else 0
        
        other_rate = float(cost_data.get('other_expenses_rate', 0))
        other_amount = sale_price * (other_rate / 100) if other_rate > 0 else 0
        
        platform_fee = float(cost_data.get('platform_fee', 6.6))
        calculated_vat = sale_price * 0.2
        
        # Profit analysis verilerini al
        if profit_analysis:
            total_expenses = profit_analysis.get('total_expenses', 0)
            net_vat = profit_analysis.get('net_vat', 0)
            profit_amount = profit_analysis.get('profit_amount', 0)
            profit_rate = profit_analysis.get('profit_rate', 0)
        else:
            total_expenses = net_vat = profit_amount = profit_rate = 0
        
        # Diğer verileri ekle
        row_data.extend([
            cargo_cost, cargo_vat,
            commission_amount, commission_vat,
            withholding_amount, platform_fee, other_amount,
            calculated_vat, total_expenses, net_vat,
            profit_amount, profit_rate
        ])
        
        # Satırı Excel'e yaz
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Dosyayı kaydet
        wb.save(filename)
        logging.info(f"Maliyet değişikliği kaydedildi: {barcode} - {username}")
        
        return True
        
    except Exception as e:
        logging.error(f"Maliyet değişikliği kayıt hatası: {str(e)}")
        return False


def get_cost_tracking_stats():
    """Maliyet takip istatistiklerini getir"""
    try:
        filename = get_yearly_excel_filename()
        if not os.path.exists(filename):
            return {
                'exists': False,
                'filename': filename,
                'total_records': 0,
                'file_size_mb': 0
            }
        
        # Dosya bilgilerini al
        file_size = os.path.getsize(filename) / (1024 * 1024)
        
        # Excel'i aç ve satır sayısını al
        df = pd.read_excel(filename)
        total_records = len(df)
        
        return {
            'exists': True,
            'filename': filename,
            'total_records': total_records,
            'file_size_mb': round(file_size, 2),
            'creation_time': datetime.fromtimestamp(os.path.getctime(filename)).strftime("%d.%m.%Y %H:%M")
        }
        
    except Exception as e:
        logging.error(f"İstatistik alma hatası: {str(e)}")
        return {
            'exists': False,
            'error': str(e)
        }